# 导入openpyxl库，用于操作Excel文件
import openpyxl
# 导入处理日期和时间的模块
from datetime import datetime, timedelta

# Load the workbook named "behavior2023-01-01.xlsx"
workbook = openpyxl.load_workbook("python数据处理与可视化/data/behavior2023-01-01.xlsx")
# Get the active worksheet
sheet = workbook.active

# Initialize index for the time column
time_column_index = None
# Iterate through columns to find the one marked as "time"
for col_num, column in enumerate(sheet.iter_cols(min_col=1, max_col=sheet.max_column), 1):
    if column[0].value == "time":
        time_column_index = col_num
        break

# If the time column is not found, print a message
if time_column_index is None:
    print("时间列未找到")  # time column not found
else:
    # Add "日期" and "时间" headers to the last columns of the sheet
    sheet.cell(1, sheet.max_column + 1, "日期")
    sheet.cell(1, sheet.max_column + 1, "时间")  # Corrected the column index for "时间"

    # Iterate through each row to convert timestamp format
    for row_num in range(2, sheet.max_row + 1):
        time_value = sheet.cell(row_num, time_column_index).value
        # Convert timestamp to datetime object
        dt_object = datetime.utcfromtimestamp(time_value / 1000.0)
        # Fill in the date in the date column
        sheet.cell(row_num, sheet.max_column - 1, dt_object.strftime("%Y-%m-%d"))
        # Fill in the time in the time column
        sheet.cell(row_num, sheet.max_column, dt_object.strftime("%H:%M:%S"))

# Save the modified workbook
workbook.save("python数据处理与可视化/csv/hotel.csv")